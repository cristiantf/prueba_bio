from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import biometrico_driver as bio
import threading
import pandas as pd
import io
from datetime import datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

# --- CONFIGURACIÓN CENTRALIZADA ---
import config

app = Flask(__name__)
app.secret_key = config.SECRET_KEY

# --- FILTROS DE PLANTILLA ---
@app.template_filter('datetimeformat')
def datetimeformat(value, format='%d/%m/%Y %H:%M'):
    if value == 'now':
        return datetime.now().strftime(format)
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except ValueError:
            return value
    if isinstance(value, datetime):
        return value.strftime(format)
    return value

# --- CONFIGURACIÓN DE FLASK-LOGIN ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = "Por favor, inicia sesión para acceder a esta página."
login_manager.login_message_category = "info"

class User(UserMixin):
    def __init__(self, id, username, rol, nombre, bio_id, acceso_puerta=0):
        self.id = id
        self.username = username
        self.rol = rol
        self.nombre = nombre
        self.bio_id = bio_id
        self.acceso_puerta = acceso_puerta

@login_manager.user_loader
def load_user(user_id):
    conn = bio.get_db_connection()
    u = conn.execute("SELECT * FROM usuarios WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    if u:
        acceso = u['acceso_puerta'] if 'acceso_puerta' in u.keys() else 0
        return User(id=u['id'], username=u['username'], rol=u['rol'], 
                    nombre=u['nombre'], bio_id=u['biometric_id'], acceso_puerta=acceso)
    return None

# --- INICIALIZACIÓN ---
bio.init_db()
threading.Thread(target=bio.iniciar_escucha_background, daemon=True).start()

# --- RUTAS PRINCIPALES ---
@app.route('/', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('admin_dashboard') if current_user.rol == 'admin' else url_for('docente_dashboard'))

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = bio.get_db_connection()
        usuario_db = conn.execute("SELECT * FROM usuarios WHERE username = ?", (username,)).fetchone()
        conn.close()
        
        # Ahora se verifica el hash de la contraseña, no el texto plano
        if usuario_db and check_password_hash(usuario_db['password'], password):
            acceso = usuario_db['acceso_puerta'] if 'acceso_puerta' in usuario_db.keys() else 0
            user_obj = User(id=usuario_db['id'], username=usuario_db['username'], rol=usuario_db['rol'], 
                            nombre=usuario_db['nombre'], bio_id=usuario_db['biometric_id'], acceso_puerta=acceso)
            login_user(user_obj)
            
            # Redirige al dashboard correspondiente
            if user_obj.rol == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('docente_dashboard'))
        else:
            flash('Usuario o contraseña incorrectos.', 'danger')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Has cerrado sesión exitosamente.', 'success')
    return redirect(url_for('login'))

# --- RUTAS DE ADMINISTRADOR ---
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.rol != 'admin':
        return redirect(url_for('docente_dashboard'))
        
    conn = bio.get_db_connection()
    docentes = [dict(row) for row in conn.execute("SELECT * FROM usuarios WHERE rol='docente' ORDER BY nombre").fetchall()]
    query_logs = '''
        SELECT l.fecha, l.usuario_id, u.nombre, l.tipo_evento, l.origen
        FROM logs l 
        LEFT JOIN usuarios u ON l.usuario_id = u.biometric_id 
        ORDER BY l.id DESC LIMIT 20
    '''
    logs = [dict(row) for row in conn.execute(query_logs).fetchall()]
    conn.close()
    
    return render_template('admin.html', docentes=docentes, logs=logs)

@app.route('/toggle_permiso/<int:id>', methods=['POST'])
@login_required
def toggle_permiso(id):
    if current_user.rol != 'admin':
        return jsonify({'success': False, 'message': 'No autorizado'}), 403
    
    try:
        data = request.get_json()
        nuevo_estado = 1 if data.get('estado') else 0
        
        conn = bio.get_db_connection()
        conn.execute("UPDATE usuarios SET acceso_puerta = ? WHERE id = ?", (nuevo_estado, id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en toggle_permiso: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/crear_docente', methods=['POST'])
@login_required
def crear_docente():
    if current_user.rol != 'admin': return redirect(url_for('login'))
    
    nombre = request.form['nombre']
    bio_id = request.form['bio_id']
    username = request.form['username']
    password = request.form['password']
    acceso = 1 if request.form.get('acceso_puerta') else 0 

    if not all([nombre, bio_id, username, password]):
        flash('Todos los campos son obligatorios.', 'warning')
        return redirect(url_for('admin_dashboard'))

    # Se hashea la contraseña antes de guardarla
    hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
    
    try:
        conn = bio.get_db_connection()
        conn.execute(
            "INSERT INTO usuarios (biometric_id, nombre, username, password, rol, acceso_puerta) VALUES (?, ?, ?, ?, 'docente', ?)",
            (bio_id, nombre, username, hashed_password, acceso)
        )
        conn.commit()
        conn.close()
        flash(f'Docente "{nombre}" creado exitosamente.', 'success')
    except Exception as e:
        flash(f'Error al crear el docente: {e}', 'danger')
        
    return redirect(url_for('admin_dashboard'))

@app.route('/eliminar_docente/<int:id>')
@login_required
def eliminar_docente(id):
    if current_user.rol != 'admin': return redirect(url_for('login'))
    
    conn = bio.get_db_connection()
    conn.execute("DELETE FROM usuarios WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    
    flash('Docente eliminado correctamente.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/editar_docente/<int:id>')
@login_required
def editar_docente(id):
    if current_user.rol != 'admin': return redirect(url_for('login'))
    
    conn = bio.get_db_connection()
    docente = conn.execute("SELECT * FROM usuarios WHERE id = ?", (id,)).fetchone()
    conn.close()
    
    if docente:
        return render_template('editar_docente.html', docente=dict(docente))
    
    flash('El docente no fue encontrado.', 'warning')
    return redirect(url_for('admin_dashboard'))

@app.route('/actualizar_docente', methods=['POST'])
@login_required
def actualizar_docente():
    if current_user.rol != 'admin': return redirect(url_for('login'))
    
    docente_id = request.form['docente_id']
    nombre = request.form['nombre']
    bio_id = request.form['bio_id']
    username = request.form['username']
    password = request.form.get('password') # Es opcional
    acceso = 1 if request.form.get('acceso_puerta') else 0
    
    conn = bio.get_db_connection()
    # Si se proporcionó una nueva contraseña, se hashea
    if password:
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        conn.execute(
            "UPDATE usuarios SET biometric_id=?, nombre=?, username=?, password=?, acceso_puerta=? WHERE id=?",
            (bio_id, nombre, username, hashed_password, acceso, docente_id)
        )
    else:
        # Si no, se actualiza sin tocar la contraseña
        conn.execute(
            "UPDATE usuarios SET biometric_id=?, nombre=?, username=?, acceso_puerta=? WHERE id=?",
            (bio_id, nombre, username, acceso, docente_id)
        )
    conn.commit()
    conn.close()
    
    flash('Docente actualizado correctamente.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/abrir_puerta')
@login_required
def admin_abrir():
    if current_user.rol != 'admin': return redirect(url_for('login'))
    
    exito, msg = bio.abrir_puerta_remota(f"Admin: {current_user.nombre}")
    flash(msg, 'success' if exito else 'error')
    
    return redirect(url_for('admin_dashboard'))

# --- API & REPORTES ---
@app.route('/api/logs')
@login_required
def api_logs():
    if current_user.rol != 'admin': return jsonify({"error": "No autorizado"}), 403
    
    conn = bio.get_db_connection()
    logs = [dict(row) for row in conn.execute("SELECT l.id, l.fecha, l.usuario_id, u.nombre, l.tipo_evento, l.origen FROM logs l LEFT JOIN usuarios u ON l.usuario_id = u.biometric_id ORDER BY l.id DESC LIMIT 20").fetchall()]
    conn.close()
    
    return jsonify(logs)

@app.route('/descargar_reporte_matricial')
@login_required
def descargar_reporte_matricial():
    if current_user.rol != 'admin': return redirect(url_for('login'))
    
    fecha_ini_str = request.args.get('fecha_inicio')
    fecha_fin_str = request.args.get('fecha_fin')
    docente_id = request.args.get('docente_id')

    try:
        end_date = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else datetime.now().date()
        start_date = datetime.strptime(fecha_ini_str, '%Y-%m-%d').date() if fecha_ini_str else end_date - timedelta(days=15)
    except ValueError:
        flash("Formato de fecha inválido. Use YYYY-MM-DD.", "danger")
        return redirect(url_for('admin_dashboard'))

    date_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]

    conn = bio.get_db_connection()
    if docente_id and docente_id != 'todos':
        users = conn.execute("SELECT * FROM usuarios WHERE biometric_id = ?", (docente_id,)).fetchall()
    else:
        users = conn.execute("SELECT * FROM usuarios WHERE rol='docente'").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Matricial"
    
    # Estilos (sin cambios)
    blue_fill = PatternFill(start_color="CCECFF", end_color="CCECFF", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    gray_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('A1:Z1')
    ws['A1'].value = f"Reporte de Asistencia ({start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')})"
    ws['A1'].fill = blue_fill
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['ID Biométrico', 'Nombre Completo', 'Departamento']
    for col, txt in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=txt)
        cell.fill = green_fill
        cell.border = thin_border
        ws.column_dimensions[chr(64+col)].width = 25

    for i, day in enumerate(date_range):
        col_idx = 4 + i
        cell = ws.cell(row=2, column=col_idx, value=day.strftime('%d/%m'))
        cell.fill = orange_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18
    
    corte_jornada = time(13, 50, 0)
    for row_idx, user in enumerate(users, 3):
        ws.cell(row=row_idx, column=1, value=user['biometric_id']).border = thin_border
        ws.cell(row=row_idx, column=2, value=user['nombre']).border = thin_border
        ws.cell(row=row_idx, column=3, value="Docencia").border = thin_border
        
        for col_offset, day in enumerate(date_range):
            day_str = day.strftime('%Y-%m-%d')
            logs_db = conn.execute("SELECT fecha FROM logs WHERE usuario_id = ? AND fecha LIKE ?", (user['biometric_id'], f"{day_str}%")).fetchall()
            
            marcas_manana, marcas_tarde = [], []
            for log in logs_db:
                try:
                    ts = datetime.fromisoformat(log['fecha'])
                    if ts.time() <= corte_jornada:
                        marcas_manana.append(ts)
                    else:
                        marcas_tarde.append(ts)
                except (ValueError, TypeError):
                    continue

            text_manana = f"{min(marcas_manana).strftime('%H:%M')}-{max(marcas_manana).strftime('%H:%M')}" if len(marcas_manana) > 1 else (marcas_manana[0].strftime('%H:%M') if marcas_manana else "--:--")
            text_tarde = f"{min(marcas_tarde).strftime('%H:%M')}-{max(marcas_tarde).strftime('%H:%M')}" if len(marcas_tarde) > 1 else (marcas_tarde[0].strftime('%H:%M') if marcas_tarde else "--:--")
            
            final_text = f"Mañana: {text_manana}\nTarde: {text_tarde}".strip()
            
            cell = ws.cell(row=row_idx, column=4 + col_offset, value=final_text)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            if row_idx % 2 != 0: cell.fill = gray_fill
    
    conn.close()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="Reporte_Asistencia.xlsx", as_attachment=True)

# --- RUTAS DE DOCENTE ---
@app.route('/docente')
@login_required
def docente_dashboard():
    conn = bio.get_db_connection()
    logs = [dict(row) for row in conn.execute("SELECT * FROM logs WHERE usuario_id = ? ORDER BY id DESC LIMIT 10", (current_user.bio_id,)).fetchall()]
    conn.close()
    return render_template('docente.html', logs=logs)

@app.route('/docente/abrir_puerta')
@login_required
def docente_abrir():
    if current_user.rol == 'docente' and current_user.acceso_puerta != 1:
        flash("⛔ No tienes permiso para abrir la puerta.", "warning")
        return redirect(url_for('docente_dashboard'))

    exito, msg = bio.abrir_puerta_remota(f"Docente: {current_user.nombre}")
    flash(msg, 'success' if exito else 'danger')
    return redirect(url_for('docente_dashboard'))

@app.route('/docente/marcar_web')
@login_required
def docente_marcar():
    bio.guardar_log("Ahora", current_user.bio_id, "ASISTENCIA WEB", "Panel Web")
    flash("Asistencia web registrada correctamente.", "success")
    return redirect(url_for('docente_dashboard'))

# --- PERFIL DE USUARIO ---
@app.route('/perfil')
@login_required
def perfil():
    return render_template('cambiar_password.html')

@app.route('/actualizar_password', methods=['POST'])
@login_required
def actualizar_password():
    current_pw = request.form['current_password']
    new_pw = request.form['new_password']
    confirm_pw = request.form['confirm_password']

    if not all([current_pw, new_pw, confirm_pw]):
        flash('Todos los campos son obligatorios.', 'warning')
        return redirect(url_for('perfil'))

    if new_pw != confirm_pw:
        flash('Las contraseñas nuevas no coinciden.', 'danger')
        return redirect(url_for('perfil'))

    conn = bio.get_db_connection()
    user_db = conn.execute("SELECT password FROM usuarios WHERE id = ?", (current_user.id,)).fetchone()

    if user_db and check_password_hash(user_db['password'], current_pw):
        new_hashed_password = generate_password_hash(new_pw, method='pbkdf2:sha256')
        conn.execute("UPDATE usuarios SET password = ? WHERE id = ?", (new_hashed_password, current_user.id))
        conn.commit()
        conn.close()
        flash('Contraseña actualizada exitosamente.', 'success')
        return redirect(url_for('admin_dashboard' if current_user.rol == 'admin' else 'docente_dashboard'))
    else:
        conn.close()
        flash('La contraseña actual es incorrecta.', 'danger')
        return redirect(url_for('perfil'))

if __name__ == '__main__':
    # Se recomienda desactivar el modo debug en producción
    print("--- SERVIDOR REINICIADO CORRECTAMENTE ---")
    app.run(host='0.0.0.0', port=5000, debug=True)